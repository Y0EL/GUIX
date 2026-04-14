from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from orchestration.analisis_kasus import PenghasilArtefakAnalisa
from orchestration.schema import AktorSindikat, DossierSindikat, RelasiSindikat


def _buat_dossier() -> DossierSindikat:
    return DossierSindikat(
        id_kasus="kasus-uji",
        judul_kasus="Kasus Uji Keterbacaan",
        ringkasan_eksekutif="Ringkasan eksekutif untuk pengujian artefak dokumen.",
        indikasi_sindikat=True,
        confidence=0.82,
        alasan_utama=["Ada pola transfer berulang.", "Aktor yang sama muncul lintas kanal."],
        pola_koordinasi=["Koordinasi berbasis kanal pesan instan."],
        aktor_inti=[
            AktorSindikat(
                id_profil="prof-1",
                nama="Adi Supriyanto",
                peran="Koordinator",
                alasan="Muncul sebagai simpul komunikasi utama.",
                confidence=0.9,
            )
        ],
        relasi_kunci=[
            RelasiSindikat(
                sumber="Adi Supriyanto",
                target="Budi Raharja",
                jenis_relasi="mentransfer",
                alasan="Terdapat histori transfer yang konsisten.",
                confidence=0.86,
            )
        ],
        bukti_utama=["Transfer bernilai tinggi terjadi berulang."],
        bukti_lemah=["Belum ada konfirmasi tatap muka langsung."],
        rekomendasi_lanjutan=["Lakukan korelasi dengan lokasi historis."],
        narasi_analisis="Narasi analisis pengujian untuk memastikan dokumen terstruktur.",
    )


def _buat_bundel() -> dict:
    return {
        "laporan": [{"id_laporan": "lap-1"}],
        "skor_risiko": [{"id_profil": "prof-1", "skor": 87}],
        "transaksi": [
            {
                "id_transaksi": "trx-1",
                "nama_sumber": "Adi Supriyanto",
                "nama_tujuan": "Budi Raharja",
                "jumlah_idr": 15000000,
                "kanal": "transfer-bank",
                "timestamp": "2026-04-14T12:00:00+07:00",
            }
        ],
        "kampanye": [],
        "profil": [],
        "lokasi": [],
        "postingan": [],
        "graf": {"nodes": [], "edges": []},
    }


def test_artefak_dokumen_memakai_struktur_bergaya(tmp_path: Path) -> None:
    penghasil = PenghasilArtefakAnalisa(tmp_path)
    hasil = penghasil.simpan_semua("kasus-uji", _buat_dossier(), _buat_bundel())
    direktori = Path(hasil["direktori"])

    workbook = load_workbook(next(direktori.glob("*.xlsx")))
    assert workbook.sheetnames == ["Dashboard", "Aktor Inti", "Relasi Kunci", "Transaksi", "Bukti & Rekom"]
    assert workbook["Dashboard"]["A10"].value == "Ringkasan Eksekutif"
    assert "Kasus Uji Keterbacaan" in str(workbook["Dashboard"]["A1"].value)

    dokumen = Document(next(direktori.glob("*.docx")))
    paragraf = [paragraf.text for paragraf in dokumen.paragraphs if paragraf.text.strip()]
    assert "Ringkasan Eksekutif" in paragraf
    assert "Bukti dan Rekomendasi" in paragraf

    pdf = next(direktori.glob("*.pdf"))
    assert pdf.read_bytes().startswith(b"%PDF")
    assert pdf.stat().st_size > 1000
